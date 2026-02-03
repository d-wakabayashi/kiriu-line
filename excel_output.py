"""
KIRIU ライン負荷最適化システム - Excel出力モジュール
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import DISC_LINES, DEFAULT_CAPACITIES, MONTHS, OUTPUT_DIR
from model import OptimizationResult
from data_loader import PartSpec


def create_styles():
    """共通スタイルを定義"""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')

    subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    number_format = '#,##0'
    percent_format = '0.0%'

    warning_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_font_white': header_font_white,
        'subheader_fill': subheader_fill,
        'thin_border': thin_border,
        'number_format': number_format,
        'percent_format': percent_format,
        'warning_fill': warning_fill,
    }


def _get_monthly_capacity(capacities: dict[str, int | list[int]], line: str, month: int) -> int:
    """月別能力を取得するヘルパー関数"""
    cap = capacities.get(line, 0)
    if isinstance(cap, list):
        return cap[month] if month < len(cap) else cap[-1]
    return cap


def _normalize_capacities(capacities: dict[str, int | list[int]]) -> dict[str, list[int]]:
    """能力を月別形式に正規化"""
    from config import DISC_LINES, DEFAULT_CAPACITIES
    result = {}
    for line in DISC_LINES:
        cap = capacities.get(line, DEFAULT_CAPACITIES.get(line, 0))
        if isinstance(cap, list):
            result[line] = (cap + [cap[-1]] * 12)[:12] if cap else [0] * 12
        else:
            result[line] = [cap] * 12
    return result


def create_summary_sheet(wb: Workbook, result: OptimizationResult, capacities: dict[str, int | list[int]]):
    """サマリーシートを作成"""
    ws = wb.active
    ws.title = 'サマリー'
    styles = create_styles()

    # タイトル
    ws['A1'] = 'KIRIU ライン負荷最適化結果'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    # 基本情報
    ws['A3'] = '最適化ステータス'
    ws['B3'] = result.status
    ws['A4'] = '目的関数値'
    ws['B4'] = result.objective_value if result.objective_value else 'N/A'
    ws['A5'] = '実行時間（秒）'
    ws['B5'] = f'{result.solve_time:.2f}'

    # 未割当サマリー
    total_unmet = 0
    if result.unmet_demand:
        for monthly in result.unmet_demand.values():
            total_unmet += sum(monthly)
    ws['A6'] = '未割当数量合計'
    ws['B6'] = total_unmet
    ws['B6'].number_format = '#,##0'
    if total_unmet > 0:
        ws['B6'].fill = styles['warning_fill']

    # ライン別負荷サマリー（行9から）
    ws['A9'] = 'ライン別負荷サマリー'
    ws['A9'].font = Font(bold=True, size=12)

    headers = ['ライン', '平均能力', '平均負荷', '負荷率', '最大負荷', '最大月']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = styles['header_font_white']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center')

    # 月別能力を正規化
    monthly_caps = _normalize_capacities(capacities)

    row = 11
    for line in DISC_LINES:
        loads = result.line_loads.get(line, [0] * 12)
        line_caps = monthly_caps.get(line, [0] * 12)
        avg_cap = sum(line_caps) / 12
        avg_load = sum(loads) / 12
        avg_rate = avg_load / avg_cap if avg_cap > 0 else 0
        max_load = max(loads)
        max_month_idx = loads.index(max_load) if loads else 0
        max_month = MONTHS[max_month_idx]

        ws.cell(row=row, column=1, value=line).border = styles['thin_border']
        ws.cell(row=row, column=2, value=int(avg_cap)).border = styles['thin_border']
        ws.cell(row=row, column=2).number_format = styles['number_format']
        ws.cell(row=row, column=3, value=int(avg_load)).border = styles['thin_border']
        ws.cell(row=row, column=3).number_format = styles['number_format']

        rate_cell = ws.cell(row=row, column=4, value=avg_rate)
        rate_cell.border = styles['thin_border']
        rate_cell.number_format = styles['percent_format']
        if avg_rate > 1.0:
            rate_cell.fill = styles['warning_fill']

        ws.cell(row=row, column=5, value=max_load).border = styles['thin_border']
        ws.cell(row=row, column=5).number_format = styles['number_format']
        ws.cell(row=row, column=6, value=max_month).border = styles['thin_border']

        row += 1

    # 列幅調整
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10


def create_line_capacity_sheet(wb: Workbook, result: OptimizationResult, capacities: dict[str, int | list[int]]):
    """ライン別月別キャパシティ・負荷シートを作成"""
    ws = wb.create_sheet('ライン別負荷')
    styles = create_styles()

    # 月別能力を正規化
    monthly_caps = _normalize_capacities(capacities)

    # ヘッダー行
    ws['A1'] = 'ライン'
    ws['B1'] = '項目'
    for col, month in enumerate(MONTHS, start=3):
        ws.cell(row=1, column=col, value=month)
    ws.cell(row=1, column=15, value='年間計')
    ws.cell(row=1, column=16, value='平均')

    # ヘッダースタイル
    for col in range(1, 17):
        cell = ws.cell(row=1, column=col)
        cell.font = styles['header_font_white']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for line in DISC_LINES:
        line_caps = monthly_caps.get(line, [0] * 12)
        loads = result.line_loads.get(line, [0] * 12)

        # キャパシティ行（月別表示）
        ws.cell(row=row, column=1, value=line).border = styles['thin_border']
        ws.cell(row=row, column=2, value='キャパシティ').border = styles['thin_border']
        ws.cell(row=row, column=2).fill = styles['subheader_fill']
        for col, month_cap in enumerate(line_caps, start=3):
            cell = ws.cell(row=row, column=col, value=month_cap)
            cell.border = styles['thin_border']
            cell.number_format = styles['number_format']
        ws.cell(row=row, column=15, value=sum(line_caps)).border = styles['thin_border']
        ws.cell(row=row, column=15).number_format = styles['number_format']
        ws.cell(row=row, column=16, value=int(sum(line_caps)/12)).border = styles['thin_border']
        ws.cell(row=row, column=16).number_format = styles['number_format']
        row += 1

        # 負荷（生産数）行
        ws.cell(row=row, column=1, value='').border = styles['thin_border']
        ws.cell(row=row, column=2, value='生産数').border = styles['thin_border']
        for col, (load, cap) in enumerate(zip(loads, line_caps), start=3):
            cell = ws.cell(row=row, column=col, value=load)
            cell.border = styles['thin_border']
            cell.number_format = styles['number_format']
            if load > cap:
                cell.fill = styles['warning_fill']
        ws.cell(row=row, column=15, value=sum(loads)).border = styles['thin_border']
        ws.cell(row=row, column=15).number_format = styles['number_format']
        ws.cell(row=row, column=16, value=int(sum(loads)/12)).border = styles['thin_border']
        ws.cell(row=row, column=16).number_format = styles['number_format']
        row += 1

        # 負荷率行（月別能力で計算）
        ws.cell(row=row, column=1, value='').border = styles['thin_border']
        ws.cell(row=row, column=2, value='負荷率').border = styles['thin_border']
        for col, (load, cap) in enumerate(zip(loads, line_caps), start=3):
            rate = load / cap if cap > 0 else 0
            cell = ws.cell(row=row, column=col, value=rate)
            cell.border = styles['thin_border']
            cell.number_format = styles['percent_format']
            if rate > 1.0:
                cell.fill = styles['warning_fill']
        total_cap = sum(line_caps)
        avg_rate = sum(loads) / total_cap if total_cap > 0 else 0
        ws.cell(row=row, column=15, value='').border = styles['thin_border']
        ws.cell(row=row, column=16, value=avg_rate).border = styles['thin_border']
        ws.cell(row=row, column=16).number_format = styles['percent_format']
        row += 1

        # 空行
        row += 1

    # 列幅調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    for col in range(3, 17):
        ws.column_dimensions[get_column_letter(col)].width = 10


def create_part_allocation_sheet(
    wb: Workbook,
    result: OptimizationResult,
    specs: dict[str, PartSpec],
):
    """部品別生産割当シートを作成"""
    ws = wb.create_sheet('部品別割当')
    styles = create_styles()

    # ヘッダー行
    headers = ['部品番号', '部品名', 'メインライン', '割当ライン'] + MONTHS + ['年間計']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles['header_font_white']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for part_num in sorted(result.allocation.keys()):
        part_data = result.allocation[part_num]
        spec = specs.get(part_num)
        part_name = spec.part_name if spec else ''
        main_line = spec.main_line if spec else ''

        for line, monthly in part_data.items():
            if sum(monthly) == 0:
                continue

            ws.cell(row=row, column=1, value=part_num).border = styles['thin_border']
            ws.cell(row=row, column=2, value=part_name).border = styles['thin_border']
            ws.cell(row=row, column=3, value=main_line).border = styles['thin_border']

            line_cell = ws.cell(row=row, column=4, value=line)
            line_cell.border = styles['thin_border']
            # サブラインの場合は色付け
            if line != main_line:
                line_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

            for col, qty in enumerate(monthly, start=5):
                cell = ws.cell(row=row, column=col, value=qty)
                cell.border = styles['thin_border']
                cell.number_format = styles['number_format']

            ws.cell(row=row, column=17, value=sum(monthly)).border = styles['thin_border']
            ws.cell(row=row, column=17).number_format = styles['number_format']

            row += 1

    # 列幅調整
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    for col in range(5, 18):
        ws.column_dimensions[get_column_letter(col)].width = 10


def create_line_detail_sheets(
    wb: Workbook,
    result: OptimizationResult,
    specs: dict[str, PartSpec],
    capacities: dict[str, int | list[int]],
):
    """ライン別詳細シートを作成（各ラインごとに1シート）"""
    styles = create_styles()

    # 月別能力を正規化
    monthly_caps = _normalize_capacities(capacities)

    for line in DISC_LINES:
        ws = wb.create_sheet(f'L{line}')
        line_caps = monthly_caps.get(line, [0] * 12)
        avg_cap = int(sum(line_caps) / 12)

        # タイトル
        ws['A1'] = f'ライン {line} 生産計画'
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:N1')

        ws['A2'] = f'平均月間キャパシティ: {avg_cap:,}'

        # ヘッダー行
        headers = ['部品番号', '割当区分'] + MONTHS + ['年間計']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = styles['header_font_white']
            cell.fill = styles['header_fill']
            cell.border = styles['thin_border']
            cell.alignment = Alignment(horizontal='center')

        row = 5
        for part_num, part_data in sorted(result.allocation.items()):
            if line not in part_data:
                continue

            monthly = part_data[line]
            if sum(monthly) == 0:
                continue

            spec = specs.get(part_num)
            main_line = spec.main_line if spec else ''
            is_sub = line != main_line

            ws.cell(row=row, column=1, value=part_num).border = styles['thin_border']

            alloc_cell = ws.cell(row=row, column=2, value='サブ' if is_sub else 'メイン')
            alloc_cell.border = styles['thin_border']
            if is_sub:
                alloc_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

            for col, qty in enumerate(monthly, start=3):
                cell = ws.cell(row=row, column=col, value=qty)
                cell.border = styles['thin_border']
                cell.number_format = styles['number_format']

            ws.cell(row=row, column=15, value=sum(monthly)).border = styles['thin_border']
            ws.cell(row=row, column=15).number_format = styles['number_format']

            row += 1

        # 合計行
        row += 1
        loads = result.line_loads.get(line, [0] * 12)

        ws.cell(row=row, column=1, value='合計').font = Font(bold=True)
        ws.cell(row=row, column=1).border = styles['thin_border']
        ws.cell(row=row, column=2, value='').border = styles['thin_border']
        for col, (load, cap) in enumerate(zip(loads, line_caps), start=3):
            cell = ws.cell(row=row, column=col, value=load)
            cell.border = styles['thin_border']
            cell.number_format = styles['number_format']
            cell.font = Font(bold=True)
            if load > cap:
                cell.fill = styles['warning_fill']
        ws.cell(row=row, column=15, value=sum(loads)).border = styles['thin_border']
        ws.cell(row=row, column=15).number_format = styles['number_format']
        ws.cell(row=row, column=15).font = Font(bold=True)

        # キャパシティ行（月別表示）
        row += 1
        ws.cell(row=row, column=1, value='キャパシティ').font = Font(bold=True)
        ws.cell(row=row, column=1).border = styles['thin_border']
        ws.cell(row=row, column=2, value='').border = styles['thin_border']
        for col, cap in enumerate(line_caps, start=3):
            cell = ws.cell(row=row, column=col, value=cap)
            cell.border = styles['thin_border']
            cell.number_format = styles['number_format']
        ws.cell(row=row, column=15, value=sum(line_caps)).border = styles['thin_border']
        ws.cell(row=row, column=15).number_format = styles['number_format']

        # 負荷率行（月別能力で計算）
        row += 1
        ws.cell(row=row, column=1, value='負荷率').font = Font(bold=True)
        ws.cell(row=row, column=1).border = styles['thin_border']
        ws.cell(row=row, column=2, value='').border = styles['thin_border']
        for col, (load, cap) in enumerate(zip(loads, line_caps), start=3):
            rate = load / cap if cap > 0 else 0
            cell = ws.cell(row=row, column=col, value=rate)
            cell.border = styles['thin_border']
            cell.number_format = styles['percent_format']
            if rate > 1.0:
                cell.fill = styles['warning_fill']

        # 列幅調整
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 10
        for col in range(3, 16):
            ws.column_dimensions[get_column_letter(col)].width = 10


def create_unmet_demand_sheet(
    wb: Workbook,
    result: OptimizationResult,
    specs: dict[str, PartSpec],
):
    """未割当（ライン能力超過により生産できなかった分）シートを作成"""
    ws = wb.create_sheet('未割当')
    styles = create_styles()

    # タイトル
    ws['A1'] = '未割当一覧（ライン能力制約により生産できなかった数量）'
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:P1')

    # 未割当がない場合
    if not result.unmet_demand:
        ws['A3'] = '未割当はありません。全ての需要がライン能力内で充足されました。'
        return

    # ヘッダー行
    headers = ['部品番号', '部品名', 'メインライン'] + MONTHS + ['年間計']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = styles['header_font_white']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center')

    row = 4
    total_unmet = 0
    for part_num in sorted(result.unmet_demand.keys()):
        monthly_unmet = result.unmet_demand[part_num]
        if sum(monthly_unmet) == 0:
            continue

        spec = specs.get(part_num)
        part_name = spec.part_name if spec else ''
        main_line = spec.main_line if spec else ''

        ws.cell(row=row, column=1, value=part_num).border = styles['thin_border']
        ws.cell(row=row, column=2, value=part_name).border = styles['thin_border']
        ws.cell(row=row, column=3, value=main_line).border = styles['thin_border']

        for col, qty in enumerate(monthly_unmet, start=4):
            cell = ws.cell(row=row, column=col, value=qty)
            cell.border = styles['thin_border']
            cell.number_format = styles['number_format']
            if qty > 0:
                cell.fill = styles['warning_fill']

        annual_total = sum(monthly_unmet)
        total_unmet += annual_total
        ws.cell(row=row, column=16, value=annual_total).border = styles['thin_border']
        ws.cell(row=row, column=16).number_format = styles['number_format']
        ws.cell(row=row, column=16).fill = styles['warning_fill']

        row += 1

    # 合計行
    row += 1
    ws.cell(row=row, column=1, value='合計').font = Font(bold=True)
    ws.cell(row=row, column=1).border = styles['thin_border']
    for col in range(2, 16):
        ws.cell(row=row, column=col, value='').border = styles['thin_border']
    ws.cell(row=row, column=16, value=total_unmet).border = styles['thin_border']
    ws.cell(row=row, column=16).number_format = styles['number_format']
    ws.cell(row=row, column=16).font = Font(bold=True)
    ws.cell(row=row, column=16).fill = styles['warning_fill']

    # 列幅調整
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    for col in range(4, 17):
        ws.column_dimensions[get_column_letter(col)].width = 10


def export_to_excel(
    result: OptimizationResult,
    specs: dict[str, PartSpec],
    capacities: dict[str, int | list[int]] | None = None,
    output_path: str | None = None,
) -> str:
    """
    最適化結果をExcelファイルに出力

    Args:
        result: 最適化結果
        specs: 部品仕様
        capacities: ライン能力（月別可、省略時はデフォルト値）
        output_path: 出力パス（省略時はoutputディレクトリ）

    Returns:
        出力ファイルパス
    """
    caps = capacities or DEFAULT_CAPACITIES

    wb = Workbook()

    # 各シートを作成
    create_summary_sheet(wb, result, caps)
    create_line_capacity_sheet(wb, result, caps)
    create_part_allocation_sheet(wb, result, specs)
    create_unmet_demand_sheet(wb, result, specs)  # 未割当シートを追加
    create_line_detail_sheets(wb, result, specs, caps)

    # 保存
    if output_path:
        save_path = output_path
    else:
        output_dir = Path(OUTPUT_DIR)
        output_dir.mkdir(parents=True, exist_ok=True)
        save_path = str(output_dir / 'optimization_result.xlsx')

    wb.save(save_path)
    print(f"Excel保存: {save_path}")

    return save_path


if __name__ == '__main__':
    # テスト用のダミーデータで動作確認
    from model import OptimizationResult

    dummy_loads = {
        '4915': [65000, 70000, 68000, 72000, 65000, 60000, 55000, 70000, 75000, 68000, 62000, 58000],
        '4919': [75000, 80000, 78000, 82000, 75000, 70000, 65000, 80000, 85000, 78000, 72000, 68000],
        '4927': [35000, 38000, 40000, 42000, 38000, 35000, 32000, 40000, 43000, 38000, 35000, 32000],
        '4928': [35000, 38000, 40000, 42000, 38000, 35000, 32000, 40000, 43000, 38000, 35000, 32000],
        '4934': [45000, 48000, 50000, 52000, 48000, 45000, 42000, 50000, 53000, 48000, 45000, 42000],
        '4935': [80000, 85000, 83000, 87000, 80000, 75000, 70000, 85000, 90000, 83000, 77000, 73000],
        '4945': [45000, 48000, 50000, 52000, 48000, 45000, 42000, 50000, 53000, 48000, 45000, 42000],
        '4G01': [45000, 48000, 50000, 52000, 48000, 45000, 42000, 50000, 53000, 48000, 45000, 42000],
        '4J01': [8000, 9000, 10000, 11000, 9000, 8000, 7000, 10000, 12000, 9000, 8000, 7000],
    }

    dummy_allocation = {
        'PART001': {'4915': [5000] * 12},
        'PART002': {'4919': [8000] * 12},
    }

    dummy_result = OptimizationResult(
        status='OPTIMAL',
        objective_value=12345.0,
        allocation=dummy_allocation,
        line_loads=dummy_loads,
        overflow={line: [0] * 12 for line in DISC_LINES},
        sub_line_usage={},
        solve_time=30.5,
    )

    from data_loader import PartSpec
    dummy_specs = {
        'PART001': PartSpec('PART001', 'テスト部品1', '4915', None, None),
        'PART002': PartSpec('PART002', 'テスト部品2', '4919', '4915', None),
    }

    export_to_excel(dummy_result, dummy_specs)
