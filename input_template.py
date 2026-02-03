"""
KIRIU ライン負荷最適化システム - 入力テンプレートモジュール

入力データをExcelテンプレートで管理するための機能を提供
"""

from pathlib import Path
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

from config import DISC_LINES, DEFAULT_CAPACITIES, MONTHS


@dataclass
class InputConfig:
    """入力設定"""
    # ファイルパス
    spec_file: str
    plan_file: str
    plan_sheet: str

    # ソルバー設定
    time_limit: int

    # ライン能力（月別対応: {ライン: [月別能力]} または {ライン: 固定能力}）
    capacities: dict[str, list[int] | int]

    # 出力設定
    output_dir: str
    output_to_gdrive: bool
    gdrive_folder_id: str
    send_email: bool
    email_to: str
    email_subject: str


def create_styles():
    """共通スタイルを定義"""
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    input_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        'header_fill': header_fill,
        'header_font': header_font,
        'subheader_fill': subheader_fill,
        'input_fill': input_fill,
        'thin_border': thin_border,
    }


def create_settings_sheet(wb: Workbook):
    """設定シートを作成"""
    ws = wb.active
    ws.title = '設定'
    styles = create_styles()

    # タイトル
    ws['A1'] = 'KIRIU ライン負荷最適化 - 入力設定'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    # ファイル設定セクション
    ws['A3'] = '【入力ファイル設定】'
    ws['A3'].font = Font(bold=True, size=12)

    settings = [
        ('A4', '設備仕様ファイル', 'B4', '', '設備仕様が記載されたExcelファイルのパス'),
        ('A5', '生産計画ファイル', 'B5', '', '生産計画が記載されたExcel/マクロファイルのパス'),
        ('A6', '生産計画シート名', 'B6', '赤堀分工場2029上期', '読み込むシート名'),
    ]

    for label_cell, label, value_cell, default, comment_text in settings:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[label_cell].border = styles['thin_border']

        ws[value_cell] = default
        ws[value_cell].fill = styles['input_fill']
        ws[value_cell].border = styles['thin_border']
        ws[value_cell].comment = Comment(comment_text, 'System')

    # ソルバー設定セクション
    ws['A8'] = '【ソルバー設定】'
    ws['A8'].font = Font(bold=True, size=12)

    ws['A9'] = '制限時間（秒）'
    ws['A9'].font = Font(bold=True)
    ws['A9'].border = styles['thin_border']
    ws['B9'] = 300
    ws['B9'].fill = styles['input_fill']
    ws['B9'].border = styles['thin_border']
    ws['B9'].comment = Comment('最適化の最大実行時間（秒）。通常は300秒で十分', 'System')

    # 出力設定セクション
    ws['A11'] = '【出力設定】'
    ws['A11'].font = Font(bold=True, size=12)

    ws['A12'] = '出力ディレクトリ'
    ws['A12'].font = Font(bold=True)
    ws['A12'].border = styles['thin_border']
    ws['B12'] = './output'
    ws['B12'].fill = styles['input_fill']
    ws['B12'].border = styles['thin_border']

    ws['A13'] = 'Google Drive出力'
    ws['A13'].font = Font(bold=True)
    ws['A13'].border = styles['thin_border']
    ws['B13'] = 'OFF'
    ws['B13'].fill = styles['input_fill']
    ws['B13'].border = styles['thin_border']
    # ドロップダウン
    dv = DataValidation(type='list', formula1='"ON,OFF"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws['B13'])

    ws['A14'] = 'Google DriveフォルダID'
    ws['A14'].font = Font(bold=True)
    ws['A14'].border = styles['thin_border']
    ws['B14'] = ''
    ws['B14'].fill = styles['input_fill']
    ws['B14'].border = styles['thin_border']
    ws['B14'].comment = Comment('Google DriveのフォルダIDを入力（URLの末尾の文字列）', 'System')

    ws['A16'] = '【メール設定】'
    ws['A16'].font = Font(bold=True, size=12)

    ws['A17'] = 'メール送信'
    ws['A17'].font = Font(bold=True)
    ws['A17'].border = styles['thin_border']
    ws['B17'] = 'OFF'
    ws['B17'].fill = styles['input_fill']
    ws['B17'].border = styles['thin_border']
    dv2 = DataValidation(type='list', formula1='"ON,OFF"', allow_blank=False)
    ws.add_data_validation(dv2)
    dv2.add(ws['B17'])

    ws['A18'] = '送信先メールアドレス'
    ws['A18'].font = Font(bold=True)
    ws['A18'].border = styles['thin_border']
    ws['B18'] = ''
    ws['B18'].fill = styles['input_fill']
    ws['B18'].border = styles['thin_border']

    ws['A19'] = 'メール件名'
    ws['A19'].font = Font(bold=True)
    ws['A19'].border = styles['thin_border']
    ws['B19'] = 'ライン負荷最適化結果'
    ws['B19'].fill = styles['input_fill']
    ws['B19'].border = styles['thin_border']

    # 列幅調整
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def create_capacity_sheet(wb: Workbook):
    """ライン能力シートを作成"""
    ws = wb.create_sheet('ライン能力')
    styles = create_styles()

    # タイトル
    ws['A1'] = 'ライン別月間能力設定'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:N1')

    ws['A2'] = '※黄色セルに能力値を入力してください。月ごとに異なる能力を設定可能です。'
    ws['A2'].font = Font(color='666666', italic=True)

    # ヘッダー行
    headers = ['ライン'] + MONTHS + ['備考']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center')

    # ライン別能力データ
    row = 5
    for line in DISC_LINES:
        ws.cell(row=row, column=1, value=line).border = styles['thin_border']
        ws.cell(row=row, column=1).font = Font(bold=True)

        default_cap = DEFAULT_CAPACITIES.get(line, 0)
        for col in range(2, 14):  # 月別能力
            cell = ws.cell(row=row, column=col, value=default_cap)
            cell.fill = styles['input_fill']
            cell.border = styles['thin_border']
            cell.number_format = '#,##0'

        # 備考列
        ws.cell(row=row, column=14, value='').border = styles['thin_border']

        row += 1

    # 合計行
    ws.cell(row=row + 1, column=1, value='合計').font = Font(bold=True)
    for col in range(2, 14):
        cell = ws.cell(row=row + 1, column=col)
        cell.value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{row-1})'
        cell.font = Font(bold=True)
        cell.border = styles['thin_border']
        cell.number_format = '#,##0'

    # 列幅調整
    ws.column_dimensions['A'].width = 10
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 10


def create_parts_master_sheet(wb: Workbook):
    """部品マスタシート（オプション）を作成"""
    ws = wb.create_sheet('部品マスタ')
    styles = create_styles()

    # タイトル
    ws['A1'] = '部品マスタ（追加・上書き用）'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    ws['A2'] = '※設備仕様ファイルに無い部品や、ライン割当を上書きしたい場合に使用'
    ws['A2'].font = Font(color='666666', italic=True)

    # ヘッダー行
    headers = ['部品番号', '部品名', 'メインライン', 'サブ1ライン', 'サブ2ライン', '備考']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center')

    # 入力行のテンプレート（10行分）
    for row in range(5, 15):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col, value='')
            cell.fill = styles['input_fill']
            cell.border = styles['thin_border']

    # ライン選択用のドロップダウン
    line_list = ','.join(DISC_LINES)
    dv = DataValidation(type='list', formula1=f'"{line_list}"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(5, 15):
        dv.add(ws.cell(row=row, column=3))
        dv.add(ws.cell(row=row, column=4))
        dv.add(ws.cell(row=row, column=5))

    # 列幅調整
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20


def generate_input_template(output_path: str = 'input_template.xlsx') -> str:
    """
    入力テンプレートExcelを生成

    Args:
        output_path: 出力ファイルパス

    Returns:
        生成したファイルのパス
    """
    wb = Workbook()

    create_settings_sheet(wb)
    create_capacity_sheet(wb)
    create_parts_master_sheet(wb)

    wb.save(output_path)
    print(f"入力テンプレート生成: {output_path}")

    return output_path


def load_input_config(template_path: str) -> InputConfig:
    """
    入力テンプレートから設定を読み込む

    Args:
        template_path: テンプレートファイルパス

    Returns:
        InputConfig
    """
    wb = load_workbook(template_path, data_only=True)

    # 設定シートから読み込み
    ws_settings = wb['設定']

    spec_file = ws_settings['B4'].value or ''
    plan_file = ws_settings['B5'].value or ''
    plan_sheet = ws_settings['B6'].value or ''
    time_limit = int(ws_settings['B9'].value or 300)
    output_dir = ws_settings['B12'].value or './output'
    output_to_gdrive = str(ws_settings['B13'].value).upper() == 'ON'
    gdrive_folder_id = ws_settings['B14'].value or ''
    send_email = str(ws_settings['B17'].value).upper() == 'ON'
    email_to = ws_settings['B18'].value or ''
    email_subject = ws_settings['B19'].value or 'ライン負荷最適化結果'

    # ライン能力シートから読み込み
    ws_capacity = wb['ライン能力']
    capacities = {}

    for row in range(5, 5 + len(DISC_LINES)):
        line = ws_capacity.cell(row=row, column=1).value
        if line and line in DISC_LINES:
            monthly_caps = []
            for col in range(2, 14):
                val = ws_capacity.cell(row=row, column=col).value
                monthly_caps.append(int(val) if val else 0)

            # 全月同じなら固定値として保存
            if len(set(monthly_caps)) == 1:
                capacities[line] = monthly_caps[0]
            else:
                capacities[line] = monthly_caps

    wb.close()

    return InputConfig(
        spec_file=spec_file,
        plan_file=plan_file,
        plan_sheet=plan_sheet,
        time_limit=time_limit,
        capacities=capacities,
        output_dir=output_dir,
        output_to_gdrive=output_to_gdrive,
        gdrive_folder_id=gdrive_folder_id,
        send_email=send_email,
        email_to=email_to,
        email_subject=email_subject,
    )


def load_parts_master(template_path: str) -> dict:
    """
    部品マスタシートから追加部品情報を読み込む

    Args:
        template_path: テンプレートファイルパス

    Returns:
        dict[部品番号, {part_name, main_line, sub1_line, sub2_line}]
    """
    from data_loader import normalize_part_number, normalize_line_name

    wb = load_workbook(template_path, data_only=True)
    ws = wb['部品マスタ']

    parts = {}
    for row in range(5, ws.max_row + 1):
        part_num_raw = ws.cell(row=row, column=1).value
        if not part_num_raw:
            continue

        part_num = normalize_part_number(str(part_num_raw))
        if not part_num:
            continue

        part_name = ws.cell(row=row, column=2).value or ''
        main_line = normalize_line_name(str(ws.cell(row=row, column=3).value or ''))
        sub1_line = normalize_line_name(str(ws.cell(row=row, column=4).value or ''))
        sub2_line = normalize_line_name(str(ws.cell(row=row, column=5).value or ''))

        parts[part_num] = {
            'part_name': part_name,
            'main_line': main_line,
            'sub1_line': sub1_line,
            'sub2_line': sub2_line,
        }

    wb.close()
    return parts


def get_monthly_capacities(capacities: dict[str, list[int] | int]) -> dict[str, list[int]]:
    """
    能力設定を月別形式に変換

    Args:
        capacities: {ライン: 固定値} または {ライン: [月別値]}

    Returns:
        {ライン: [12ヶ月分の能力]}
    """
    result = {}
    for line, cap in capacities.items():
        if isinstance(cap, list):
            result[line] = cap
        else:
            result[line] = [cap] * 12
    return result


if __name__ == '__main__':
    # テンプレート生成
    generate_input_template('input_template.xlsx')

    # 読み込みテスト（ファイルがあれば）
    from pathlib import Path
    if Path('input_template.xlsx').exists():
        config = load_input_config('input_template.xlsx')
        print("\n読み込んだ設定:")
        print(f"  設備仕様ファイル: {config.spec_file}")
        print(f"  生産計画ファイル: {config.plan_file}")
        print(f"  制限時間: {config.time_limit}秒")
        print(f"  ライン能力: {config.capacities}")
